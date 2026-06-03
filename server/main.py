from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response
from pydantic import BaseModel, EmailStr
from pathlib import Path
from datetime import datetime
from zoneinfo import ZoneInfo
import openpyxl
import sqlite3
import json
import shutil

app = FastAPI()

@app.get("/")
def root():
    return JSONResponse({"status": "Vulbright contact server is running."})

@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    return Response(status_code=204)

ALLOWED_ORIGINS = [
    "http://localhost:5173",
    "http://localhost:5174",
    "http://localhost:5175",
    "http://localhost:5176",
    "http://localhost:5177",
    # Production — replace with your actual Vercel URL after deploying
    "https://vulbright-website.vercel.app/",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_methods=["POST", "GET"],
    allow_headers=["Content-Type"],
)

DATA_DIR = Path(__file__).parent / "data"
EXCEL_FILE = DATA_DIR / "contacts.xlsx"
DB_FILE = DATA_DIR / "vulbright.db"
RESUMES_DIR = DATA_DIR / "resumes"
SHEET_NAME = "Contacts"
HEADERS = ["Name", "Email", "Message", "Submitted At"]

JOBS_SEED = [
    {
        "id": 1,
        "title": "Senior Cloud Architect",
        "department": "Cloud & Infrastructure",
        "location": "Spring, TX / Remote",
        "type": "Full-time",
        "description": "Lead cloud architecture design and migration projects for enterprise clients across AWS, Azure, and GCP. You'll work directly with clients to assess their infrastructure and deliver scalable, cost-optimized solutions.",
        "requirements": json.dumps([
            "7+ years of cloud architecture experience (AWS / Azure / GCP)",
            "AWS Solutions Architect Professional or Azure Expert certification",
            "Strong hands-on experience with Terraform, Kubernetes, and CI/CD pipelines",
            "Excellent client-facing communication skills",
        ]),
    },
    {
        "id": 2,
        "title": "Machine Learning Engineer",
        "department": "AI & Data",
        "location": "Spring, TX / Remote",
        "type": "Full-time",
        "description": "Build and deploy production-grade ML models for clients across finance, healthcare, and retail. You'll work end-to-end from data preparation through MLOps deployment, collaborating closely with data scientists and software engineers.",
        "requirements": json.dumps([
            "5+ years of ML engineering experience",
            "Proficiency in Python, TensorFlow or PyTorch",
            "Experience building MLOps pipelines (MLflow, SageMaker, or Vertex AI)",
            "Strong background in NLP or computer vision preferred",
        ]),
    },
    {
        "id": 3,
        "title": "Data Engineer",
        "department": "Data Engineering",
        "location": "Remote",
        "type": "Full-time",
        "description": "Design and build scalable data pipelines for enterprise clients, migrating legacy ETL to modern cloud-native architectures. You'll ensure data quality, reliability, and observability across streaming and batch workloads.",
        "requirements": json.dumps([
            "4+ years in data engineering roles",
            "Strong experience with Apache Spark, Kafka, and Airflow",
            "Hands-on expertise with dbt, Snowflake, or Databricks",
            "Solid Python and SQL skills",
        ]),
    },
    {
        "id": 4,
        "title": "Full Stack Developer",
        "department": "Software Development",
        "location": "Spring, TX / Remote",
        "type": "Full-time",
        "description": "Develop modern web applications and enterprise software for clients ranging from startups to Fortune 500 companies. You'll work in agile sprints with a focus on clean architecture, automated testing, and CI/CD delivery.",
        "requirements": json.dumps([
            "4+ years of full-stack development experience",
            "Proficiency in React, TypeScript, and Node.js",
            "Experience with PostgreSQL or similar databases",
            "Familiarity with Docker, Kubernetes, and GitHub Actions",
        ]),
    },
    {
        "id": 5,
        "title": "Digital Transformation Consultant",
        "department": "Consulting",
        "location": "Spring, TX",
        "type": "Full-time",
        "description": "Guide enterprise clients through end-to-end digital transformation engagements — from strategy and process redesign to technology implementation and change management. You'll lead workshops, manage stakeholders, and drive measurable outcomes.",
        "requirements": json.dumps([
            "6+ years in management or technology consulting",
            "Experience with Salesforce, ServiceNow, or Microsoft 365 implementations",
            "Strong workshop facilitation and stakeholder management skills",
            "MBA or equivalent experience preferred",
        ]),
    },
]


def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    RESUMES_DIR.mkdir(parents=True, exist_ok=True)
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS jobs (
            id INTEGER PRIMARY KEY,
            title TEXT NOT NULL,
            department TEXT NOT NULL,
            location TEXT NOT NULL,
            type TEXT NOT NULL,
            description TEXT NOT NULL,
            requirements TEXT NOT NULL,
            is_active INTEGER DEFAULT 1
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS applications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id INTEGER NOT NULL,
            job_title TEXT NOT NULL,
            first_name TEXT NOT NULL,
            last_name TEXT NOT NULL,
            email TEXT NOT NULL,
            phone TEXT DEFAULT '',
            consent INTEGER DEFAULT 0,
            resume_filename TEXT DEFAULT '',
            submitted_at TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS contacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT NOT NULL,
            message TEXT NOT NULL,
            submitted_at TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS newsletter_subscriptions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL UNIQUE,
            subscribed_at TEXT NOT NULL
        )
    """)
    count = conn.execute("SELECT COUNT(*) FROM jobs").fetchone()[0]
    if count == 0:
        for job in JOBS_SEED:
            conn.execute(
                "INSERT INTO jobs (id, title, department, location, type, description, requirements) VALUES (?,?,?,?,?,?,?)",
                (job["id"], job["title"], job["department"], job["location"], job["type"], job["description"], job["requirements"]),
            )
    conn.commit()
    conn.close()


init_db()


# ── Contact form (Excel) ──────────────────────────────────────────────────────

class ContactForm(BaseModel):
    name: str
    email: EmailStr
    message: str


def get_workbook():
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else _new_sheet(wb)
    else:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
    return wb, wb[SHEET_NAME]


def _new_sheet(wb):
    ws = wb.create_sheet(SHEET_NAME)
    ws.append(HEADERS)
    return ws


@app.post("/api/contact")
def submit_contact(form: ContactForm):
    if not form.name.strip() or not form.message.strip():
        raise HTTPException(status_code=400, detail="All fields are required.")
    submitted_at = datetime.now(ZoneInfo("America/Chicago")).strftime("%Y-%m-%d %H:%M:%S %Z")
    conn = get_db()
    conn.execute(
        "INSERT INTO contacts (name, email, message, submitted_at) VALUES (?,?,?,?)",
        (form.name.strip(), form.email.strip(), form.message.strip(), submitted_at),
    )
    conn.commit()
    conn.close()
    print(f"[contact] saved: {form.name} <{form.email}>")
    return {"success": True}


class NewsletterForm(BaseModel):
    email: EmailStr


@app.post("/api/newsletter")
def subscribe_newsletter(form: NewsletterForm):
    subscribed_at = datetime.now(ZoneInfo("America/Chicago")).strftime("%Y-%m-%d %H:%M:%S %Z")
    conn = get_db()
    existing = conn.execute("SELECT id FROM newsletter_subscriptions WHERE email = ?", (form.email.strip(),)).fetchone()
    if existing:
        conn.close()
        return {"success": True, "already_subscribed": True}
    conn.execute(
        "INSERT INTO newsletter_subscriptions (email, subscribed_at) VALUES (?,?)",
        (form.email.strip(), subscribed_at),
    )
    conn.commit()
    conn.close()
    print(f"[newsletter] subscribed: {form.email}")
    return {"success": True, "already_subscribed": False}


# ── Jobs ──────────────────────────────────────────────────────────────────────

@app.get("/api/jobs")
def get_jobs():
    conn = get_db()
    rows = conn.execute("SELECT * FROM jobs WHERE is_active = 1 ORDER BY id").fetchall()
    conn.close()
    return [
        {
            "id": row["id"],
            "title": row["title"],
            "department": row["department"],
            "location": row["location"],
            "type": row["type"],
            "description": row["description"],
            "requirements": json.loads(row["requirements"]),
        }
        for row in rows
    ]


# ── Applications ──────────────────────────────────────────────────────────────

@app.post("/api/applications")
async def submit_application(
    job_id: int = Form(...),
    job_title: str = Form(...),
    first_name: str = Form(...),
    last_name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(""),
    consent: str = Form("false"),
    resume: UploadFile = File(None),
):
    submitted_at = datetime.now(ZoneInfo("America/Chicago")).strftime("%Y-%m-%d %H:%M:%S %Z")
    resume_filename = ""
    if resume and resume.filename:
        ext = Path(resume.filename).suffix.lower()
        if ext not in {".pdf", ".doc", ".docx"}:
            raise HTTPException(status_code=400, detail="Only PDF, DOC, or DOCX files are accepted.")
        safe_name = f"{job_id}_{first_name}_{last_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
        with open(RESUMES_DIR / safe_name, "wb") as f:
            shutil.copyfileobj(resume.file, f)
        resume_filename = safe_name

    conn = get_db()
    conn.execute(
        "INSERT INTO applications (job_id, job_title, first_name, last_name, email, phone, consent, resume_filename, submitted_at) VALUES (?,?,?,?,?,?,?,?,?)",
        (job_id, job_title, first_name.strip(), last_name.strip(), email.strip(), phone.strip(), 1 if consent == "true" else 0, resume_filename, submitted_at),
    )
    conn.commit()
    conn.close()
    print(f"[application] {first_name} {last_name} <{email}> → {job_title}")
    return {"success": True}
